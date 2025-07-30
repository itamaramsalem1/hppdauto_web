import pandas as pd
import openpyxl
import xlrd
from datetime import datetime
import os
import re
from difflib import get_close_matches
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, numbers
import concurrent.futures
from functools import lru_cache

def normalize_name(name):
    if not name:
        return ""
    name = str(name).lower()
    name = re.sub(r"[^a-z0-9\s]", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name

@lru_cache(maxsize=1000)
def extract_core_from_report(report_name):
    print(f"        EXTRACT DEBUG: Input='{report_name}'")
    
    if not report_name:
        print(f"        EXTRACT DEBUG: Empty input, returning ''")
        return ""
    
    report_name = str(report_name).lower()
    print(f"        EXTRACT DEBUG: After lowercase='{report_name}'")

    # Remove prefix like "Total Nursing Wrkd - " if present
    if report_name.startswith("total nursing wrkd - "):
        core = report_name[21:].strip()
        print(f"        EXTRACT DEBUG: After prefix removal='{core}'")
    else:
        core = report_name.strip()
        print(f"        EXTRACT DEBUG: No prefix to remove, core='{core}'")

    # Normalize
    core = re.sub(r"[^a-z0-9\s]", "", core)
    core = re.sub(r"\s+", " ", core).strip()
    print(f"        EXTRACT DEBUG: After normalization='{core}'")

    # Apply overrides
    overrides = {
        "dallastown": "inners creek",
        "lancaster": "abbeyville",
        "montgomeryville": "montgomery",
        "west reading": "lebanon",
        "sunbury": "sunbury"  # just to be safe
    }
    
    original_core = core
    core = overrides.get(core, core)
    if core != original_core:
        print(f"        EXTRACT DEBUG: Override applied: '{original_core}' â†’ '{core}'")
    else:
        print(f"        EXTRACT DEBUG: No override, final='{core}'")
    
    return core

def build_template_name_map(template_entries):
    return {entry["cleaned_name"]: entry["facility"] for entry in template_entries}

def match_report_to_template(report_name, template_name_map, cutoff=0.6):
    print(f"\nðŸ” MATCHING DEBUG: '{report_name}'")
    
    # Step 1: Extract core name
    core_name = extract_core_from_report(report_name)
    print(f"    Step 1 - Extracted core: '{core_name}'")
    
    # Show what's available in template map
    template_keys = list(template_name_map.keys())
    print(f"    Available template keys: {template_keys}")

    # Step 2: Try exact match
    if core_name in template_name_map:
        result = template_name_map[core_name]
        print(f"    Step 2 - âœ… EXACT MATCH: '{core_name}' â†’ '{result}'")
        return result
    else:
        print(f"    Step 2 - âŒ No exact match for '{core_name}'")

    # Step 3: Try fuzzy match with high cutoff
    match = get_close_matches(core_name, template_keys, n=1, cutoff=cutoff)
    if match:
        result = template_name_map[match[0]]
        print(f"    Step 3 - âœ… FUZZY MATCH (cutoff={cutoff}): '{core_name}' â†’ '{match[0]}' â†’ '{result}'")
        return result
    else:
        print(f"    Step 3 - âŒ No fuzzy match at cutoff {cutoff}")

    # Step 4: Try low-confidence match as fallback
    match = get_close_matches(core_name, template_keys, n=1, cutoff=0.3)
    if match:
        result = template_name_map[match[0]]
        print(f"    Step 4 - âœ… LOW-CONFIDENCE MATCH (cutoff=0.3): '{core_name}' â†’ '{match[0]}' â†’ '{result}'")
        return result
    else:
        print(f"    Step 4 - âŒ No match even at cutoff 0.3")

    print(f"    FINAL RESULT: âŒ NO MATCH FOUND")
    return None

def safe_float_conversion(value, default=0.0):
    """Safely convert a value to float"""
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (ValueError, TypeError):
        return default

def safe_cell_value(ws, cell_ref):
    """Safely get cell value"""
    try:
        return ws[cell_ref].value
    except:
        return None

def safe_xlrd_cell_value(ws, row, col):
    """Safely get cell value from xlrd worksheet"""
    try:
        if ws.nrows > row and ws.ncols > col:
            return ws.cell_value(row, col)
        return None
    except:
        return None

def is_valid_file(filename, extension):
    """Check if file is valid (not a Mac OS hidden file or corrupt)"""
    if filename.startswith('._'):
        return False
    if not filename.lower().endswith(extension):
        return False
    return True

def extract_agency_cna_rnlpn_from_sheet2(ws2):
    """
    Extract agency staffing hours for CNAs and RN+LPNs from Sheet2.
    
    Args:
        ws2: xlrd worksheet object for Sheet2
        
    Returns:
        dict: {
            'agency_cna_hours': float,
            'agency_rnlpn_hours': float,
            'agency_total_hours': float
        }
    """
    agency_cna_hours = 0.0
    agency_rnlpn_hours = 0.0
    
    current_block_type = None  # 'agency_cna', 'agency_rn', 'agency_lpn', or None
    
    # Start scanning from row 11 (index 10) downward
    for row_idx in range(10, ws2.nrows):
        try:
            # Get value from column A (index 0)
            cell_value = ws2.cell_value(row_idx, 0)
            
            # Skip empty cells
            if not cell_value or str(cell_value).strip() == "":
                continue
                
            cell_str = str(cell_value).strip().upper()
            
            # Check if this is a header row (contains forward slashes)
            if '/' in cell_str:
                # Reset current block type
                current_block_type = None
                
                # Parse the header pattern: e.g., "806/AGY/.../CNA"
                parts = cell_str.split('/')
                
                # Check if this is an agency block (contains 'AGY')
                is_agency = any('AGY' in part for part in parts)
                
                if is_agency and len(parts) > 0:
                    # Get the last part to determine staff type
                    last_part = parts[-1].strip()
                    
                    if 'CNA' in last_part:
                        current_block_type = 'agency_cna'
                    elif 'RN' in last_part:
                        current_block_type = 'agency_rn'
                    elif 'LPN' in last_part:
                        current_block_type = 'agency_lpn'
            
            else:
                # This is a data row - extract hours from column M (index 12)
                if current_block_type and ws2.ncols > 12:  # Make sure column M exists
                    try:
                        hours_value = ws2.cell_value(row_idx, 12)  # Column M
                        hours = safe_float_conversion(hours_value)
                        
                        if current_block_type == 'agency_cna':
                            agency_cna_hours += hours
                        elif current_block_type in ['agency_rn', 'agency_lpn']:
                            agency_rnlpn_hours += hours
                            
                    except (ValueError, TypeError):
                        # Skip rows with invalid hour values
                        continue
                        
        except Exception:
            # Skip problematic rows
            continue
    
    return {
        'agency_cna_hours': agency_cna_hours,
        'agency_rnlpn_hours': agency_rnlpn_hours,
        'agency_total_hours': agency_cna_hours + agency_rnlpn_hours
    }

def compute_agency_percentages(agency_data, actual_cna_hours, actual_rn_hours, actual_lpn_hours):
    """
    Compute agency staffing percentages using provided hours data.
    
    Args:
        agency_data: dict from extract_agency_cna_rnlpn_from_sheet2()
        actual_cna_hours: float - actual CNA hours from hours extraction
        actual_rn_hours: float - actual RN hours from hours extraction  
        actual_lpn_hours: float - actual LPN hours from hours extraction
        
    Returns:
        dict: {
            'actual_agency_cna_pct': float,
            'actual_agency_nurse_pct': float, 
            'actual_agency_total_pct': float,
            'actual_cna_hours': float,
            'actual_rn_hours': float,
            'actual_lpn_hours': float
        }
    """
    # Calculate total nurse hours (RN + LPN)
    actual_rnlpn_hours = actual_rn_hours + actual_lpn_hours
    actual_total_hours = actual_cna_hours + actual_rnlpn_hours
    
    # Get agency hours from the input data
    agency_cna_hours = agency_data['agency_cna_hours']
    agency_rnlpn_hours = agency_data['agency_rnlpn_hours']
    agency_total_hours = agency_data['agency_total_hours']
    
    # Calculate percentages (handle divide-by-zero safely)
    actual_agency_cna_pct = (agency_cna_hours / actual_cna_hours * 100) if actual_cna_hours > 0 else 0.0
    actual_agency_nurse_pct = (agency_rnlpn_hours / actual_rnlpn_hours * 100) if actual_rnlpn_hours > 0 else 0.0
    actual_agency_total_pct = (agency_total_hours / actual_total_hours * 100) if actual_total_hours > 0 else 0.0
    
    return {
        'actual_agency_cna_pct': round(actual_agency_cna_pct, 2),
        'actual_agency_nurse_pct': round(actual_agency_nurse_pct, 2),
        'actual_agency_total_pct': round(actual_agency_total_pct, 2),
        'actual_cna_hours': actual_cna_hours,
        'actual_rn_hours': actual_rn_hours,
        'actual_lpn_hours': actual_lpn_hours
    }


def extract_hours_by_dept_code(ws3):
    """Extract hours from column H by scanning department codes in column C, starting from row 10."""
    rn_hours = lpn_hours = cna_hours = total_hours = 0.0

    for row in range(9, ws3.nrows):  # Start from row 10 (index 9)
        try:
            code_cell = ws3.cell_value(row, 2)  # Column C (index 2)
            if not code_cell:
                continue

            code = str(code_cell).strip()
            hours = safe_float_conversion(ws3.cell_value(row, 7))  # Column H (index 7)

            if code == "3210":       # RN
                rn_hours = hours
            elif code == "3215":     # LPN
                lpn_hours = hours
            elif code == "3225":     # CNA
                cna_hours = hours

            # Look for total row
            label = str(code_cell).lower()
            if "total hours worked" in label or "grand total" in label:
                total_hours = safe_float_conversion(ws3.cell_value(row, 7))

        except Exception:
            continue

    if total_hours == 0:
        total_hours = rn_hours + lpn_hours + cna_hours

    return rn_hours, lpn_hours, cna_hours, total_hours


def process_template_file(args):
    """Process a single template file - for parallel processing"""
    filepath, filename, target_date = args
    
    if not is_valid_file(filename, ".xlsx"):
        if filename.startswith('._'):
            return None, (filename, "Mac OS hidden file, skipped")
        else:
            return None, (filename, "Not .xlsx, skipped")
    
    try:
        # Use read_only=True for speed and memory efficiency
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        return None, (filename, f"Openpyxl error: {str(e)[:100]}")

    try:
        sheet_day = str(datetime.strptime(target_date, "%Y-%m-%d").day)
        if sheet_day not in wb.sheetnames:
            wb.close()
            return None, (filename, f"No sheet named '{sheet_day}'")
            
        ws = wb[sheet_day]

        # Batch read all needed cells at once
        cell_values = {}
        cells_to_read = ["D3", "E62", "B11", "E27", "G58", "E58", "F58", "L37", "L34", "O34"]
        for cell_ref in cells_to_read:
            cell_values[cell_ref] = safe_cell_value(ws, cell_ref)

        facility_full = cell_values["D3"]
        if not facility_full:
            wb.close()
            return None, (filename, "Missing facility name in D3")
            
        cleaned_facility = normalize_name(facility_full)

        # Add back the necessary mappings (but NOT the conflicting ones)
        if "sunbury skilled nursing and rehabilitation" in cleaned_facility:
            cleaned_facility = "sunbury" 
        elif "lebanon skilled nursing and rehabilitation" in cleaned_facility:
            cleaned_facility = "lebanon"
        elif "chambersburg skilled nursing and rehabilitation" in cleaned_facility:
            cleaned_facility = "chambersburg"
        elif "pottstown skilled nursing and rehabilitation" in cleaned_facility:
            cleaned_facility = "pottstown"
        # NOTE: Do NOT add back abbeyville, inners creek, or montgomery mappings

        date_cell = cell_values["B11"]
        if not date_cell:
            wb.close()
            return None, (filename, "Missing date in B11")

        try:
            if isinstance(date_cell, datetime):
                sheet_date = date_cell.date()
            else:
                sheet_date = pd.to_datetime(date_cell).date()
        except:
            wb.close()
            return None, (filename, "Invalid date format in B11")
        
        if target_date and sheet_date != datetime.strptime(target_date, "%Y-%m-%d").date():
            wb.close()
            return None, (filename, f"Date mismatch: sheet has {sheet_date}, looking for {target_date}")

        census = safe_float_conversion(cell_values["E27"])
        if census <= 0:
            wb.close()
            return None, (filename, f"Invalid census value: {census} (census must be > 0)")

        # Calculate all values at once
        cna_hours = safe_float_conversion(cell_values["G58"])
        nurse_e_hours = safe_float_conversion(cell_values["E58"])
        nurse_f_hours = safe_float_conversion(cell_values["F58"])
        
        projected_cna_hppd = cna_hours / census
        projected_nurse_hppd = (nurse_e_hours + nurse_f_hours) / census
        projected_total_hppd = projected_cna_hppd + projected_nurse_hppd

        proj_agency_total = safe_float_conversion(cell_values["L37"]) * 100
        proj_agency_nurse = safe_float_conversion(cell_values["L34"]) * 100
        proj_agency_cna = safe_float_conversion(cell_values["O34"]) * 100

        template_entry = {
            "facility": str(facility_full),
            "cleaned_name": cleaned_facility,
            "date": sheet_date,
            "note": str(cell_values["E62"]) if cell_values["E62"] else "",
            "census": census,
            "proj_total": projected_total_hppd,
            "proj_cna": projected_cna_hppd,
            "proj_nurse": projected_nurse_hppd,
            "proj_agency_total": proj_agency_total,
            "proj_agency_cna": proj_agency_cna,
            "proj_agency_nurse": proj_agency_nurse
        }
        
        wb.close()
        return template_entry, None
        
    except Exception as e:
        try:
            wb.close()
        except:
            pass
        return None, (filename, f"Data parsing error: {str(e)[:100]}")


def process_report_file(args):
    """Process a single report file - now with robust fallback from OLD to NEW hour extraction."""
    filepath, filename, target_date, template_map = args
    print(f"\nðŸ” REPORT DEBUG: Starting {filename}")

    # Step 1: Validate file
    if not is_valid_file(filename, ".xls"):
        print(f"    âŒ Invalid file type")
        if filename.startswith('._'):
            return None, (filename, "Mac OS hidden file, skipped")
        else:
            return None, (filename, "Not .xls, skipped")

    print(f"    âœ… Valid .xls file")

    # Step 2: Open workbook
    try:
        wb = xlrd.open_workbook(filepath)
    except Exception as e:
        return None, (filename, f"Failed to open workbook: {str(e)[:50]}")

    # Step 3: Extract sheets
    if "Sheet3" not in wb.sheet_names() or "Sheet2" not in wb.sheet_names():
        return None, (filename, "Missing Sheet3 or Sheet2")
    ws3 = wb.sheet_by_name("Sheet3")
    ws2 = wb.sheet_by_name("Sheet2")

    # Step 4: Parse report date
    try:
        raw_date = ws3.cell_value(3, 1)
        if isinstance(raw_date, float):
            report_date = datetime(*xlrd.xldate_as_tuple(raw_date, wb.datemode)).date()
        else:
            report_date = pd.to_datetime(raw_date).date()
    except Exception as e:
        return None, (filename, f"Invalid date format: {str(e)[:50]}")

    if target_date:
        if report_date != datetime.strptime(target_date, "%Y-%m-%d").date():
            return None, (filename, f"Date mismatch: report has {report_date}, looking for {target_date}")

    # Step 5: Extract facility name
    try:
        report_facility = ws3.cell_value(4, 1)
        if not report_facility:
            return None, (filename, "Missing facility name")
    except Exception as e:
        return None, (filename, f"Failed to extract facility name: {str(e)[:50]}")

    # Step 6: Extract hours
    print(f"    ðŸ“Š Extracting hours data...")
    try:
        # Try old method
        try:
            old_actual_hours = safe_float_conversion(ws3.cell_value(13, 7))
            old_actual_cna_hours = safe_float_conversion(ws3.cell_value(12, 7))
            old_actual_rn_hours = safe_float_conversion(ws3.cell_value(10, 7))
            old_actual_lpn_hours = safe_float_conversion(ws3.cell_value(11, 7))
            old_total = old_actual_rn_hours + old_actual_lpn_hours + old_actual_cna_hours
        except:
            print(f"    âš ï¸ OLD method failed, using new method only")
            old_actual_hours = old_actual_cna_hours = old_actual_rn_hours = old_actual_lpn_hours = old_total = 0

        # Always attempt new method
        rn_hours, lpn_hours, cna_hours, total_hours = extract_hours_by_dept_code(ws3)
        new_total = rn_hours + lpn_hours + cna_hours

        # Select method based on new result
        if new_total > 0:
            actual_hours = total_hours
            actual_cna_hours = cna_hours
            actual_rn_hours = rn_hours
            actual_lpn_hours = lpn_hours
        else:
            actual_hours = old_actual_hours
            actual_cna_hours = old_actual_cna_hours
            actual_rn_hours = old_actual_rn_hours
            actual_lpn_hours = old_actual_lpn_hours
    except Exception as e:
        return None, (filename, f"Failed to extract hours data: {str(e)[:50]}")

    # Step 7: Agency extraction
    try:
        agency_data = extract_agency_cna_rnlpn_from_sheet2(ws2)
        agency_percentages = compute_agency_percentages(
            agency_data,
            actual_cna_hours,
            actual_rn_hours,
            actual_lpn_hours
        )
    except Exception as e:
        return None, (filename, f"Failed to extract agency data: {str(e)[:50]}")

    # Step 8: Template matching
    matched_template_name = match_report_to_template(report_facility, template_map)
    if not matched_template_name:
        return None, (filename, f"No matched facility name. Report: '{extract_core_from_report(report_facility)}'")

    # âœ… STEP 3: DEBUG TRACKING
    if matched_template_name not in comparison_debug_log:
        comparison_debug_log[matched_template_name] = {
            "Template Loaded": False,
            "Census Valid": False,
            "Report Found": True,
            "Report Loaded": True,
            "Compared": False,
            "Failure Reason": "Template missing"
        }
    else:
        comparison_debug_log[matched_template_name]["Report Found"] = True
        comparison_debug_log[matched_template_name]["Report Loaded"] = True


    # Step 9: Package result
    return {
        "filename": filename,
        "report_facility": report_facility,
        "matched_template_name": matched_template_name,
        "report_date": report_date,
        "actual_hours": actual_hours,
        "actual_cna_hours": actual_cna_hours,
        "actual_rn_lpn_hours": actual_rn_hours + actual_lpn_hours,
        "actual_agency_cna_pct": agency_percentages['actual_agency_cna_pct'],
        "actual_agency_nurse_pct": agency_percentages['actual_agency_nurse_pct'],
        "actual_agency_total_pct": agency_percentages['actual_agency_total_pct']
    }, None
comparison_debug_log = {}

def run_hppd_comparison_for_date(templates_folder, reports_folder, target_date, output_path, progress_callback=None):
    print("Starting HPPD comparison...")
    
    def progress(pct, msg):
        if progress_callback:
            progress_callback(pct, msg)
        print(f"Progress {pct}%: {msg}")

    progress(5, "Collecting template files...")

    # Collect template files
    template_files = []
    for root, _, files in os.walk(templates_folder):
        for fname in files:
            template_files.append((os.path.join(root, fname), fname, target_date))
    print(f"Found {len(template_files)} template files.\n")

    # â”€â”€â”€ PHASE 1: PROCESS TEMPLATE FILES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    template_entries = []
    skipped_templates = []
    progress(15, "Processing template files...")

    # Process templates in parallel
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        results = executor.map(process_template_file, template_files)
        
        for entry, skip_info in results:
            if entry:
                template_entries.append(entry)

                # DEBUG TRACKING (Step 2)
                facility = entry["facility"]
                comparison_debug_log[facility] = {
                    "Template Loaded": True,
                    "Census Valid": entry["census"] > 0,
                    "Report Found": False,
                    "Report Loaded": False,
                    "Compared": False,
                    "Failure Reason": None
                }
                if entry["census"] <= 0:
                    comparison_debug_log[facility]["Failure Reason"] = "Invalid census (0)"

            elif skip_info:
                skipped_templates.append(skip_info)


    print(f"Processed templates: {len(template_entries)} entries, {len(skipped_templates)} skipped\n")

    # â”€â”€â”€ PHASE 2: BUILD TEMPLATE MAP â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    progress(30, "Building template map...")
    template_map = build_template_name_map(template_entries)
    print(f"[TEMPLATE MAP] {len(template_map)} keys")
    for clean, full in template_map.items():
        print(f"  â€¢ '{clean}' â†’ '{full}'")
    print()

    # â”€â”€â”€ PHASE 3: PROCESS REPORT FILES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    progress(40, "Collecting report files...")
    report_files = []
    for root, _, files in os.walk(reports_folder):
        for fname in files:
            report_files.append((os.path.join(root, fname), fname, target_date, template_map))
    print(f"Found {len(report_files)} report files.\n")

    # Process reports and collect detailed failure information
    progress(50, "Processing report files...")
    report_data_list, skipped_reports = [], []

    # Track failure types for summary
    date_failures = []
    matching_failures = []
    file_failures = []
    sheet_failures = []
    data_failures = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as ex:
        for rep, skip in ex.map(process_report_file, report_files):
            if rep: 
                report_data_list.append(rep)
            elif skip:
                skipped_reports.append(skip)
                # Categorize the failure type
                filename, reason = skip
                if "Date mismatch" in reason:
                    date_failures.append((filename, reason))
                elif "No matched facility" in reason:
                    matching_failures.append((filename, reason))
                elif "Mac OS hidden" in reason or "Not .xls" in reason:
                    file_failures.append((filename, reason))
                elif "No Sheet" in reason:
                    sheet_failures.append((filename, reason))
                else:
                    data_failures.append((filename, reason))

    # Print comprehensive summary
    print(f"\n" + "="*80)
    print(f"ðŸ“Š COMPREHENSIVE REPORT PROCESSING SUMMARY")
    print(f"="*80)
    print(f"Total reports attempted: {len(report_files)}")
    print(f"âœ… Successfully processed: {len(report_data_list)}")
    print(f"âŒ Total skipped: {len(skipped_reports)}")
    print(f"")
    print(f"FAILURE BREAKDOWN:")
    print(f"ðŸ“… Date mismatches: {len(date_failures)}")
    print(f"ðŸ”— Template matching failures: {len(matching_failures)}")
    print(f"ðŸ“ File issues (hidden/wrong extension): {len(file_failures)}")
    print(f"ðŸ“‹ Missing sheets: {len(sheet_failures)}")
    print(f"ðŸ“Š Data extraction issues: {len(data_failures)}")
    print(f"")

    # Show specific examples of each failure type
    if date_failures:
        print(f"ðŸ“… DATE FAILURE EXAMPLES:")
        for filename, reason in date_failures[:3]:
            print(f"  â€¢ {filename}: {reason}")
        if len(date_failures) > 3:
            print(f"  ... and {len(date_failures) - 3} more")
        print()

    if matching_failures:
        print(f"ðŸ”— MATCHING FAILURE EXAMPLES:")
        for filename, reason in matching_failures[:5]:
            print(f"  â€¢ {filename}: {reason}")
        if len(matching_failures) > 5:
            print(f"  ... and {len(matching_failures) - 5} more")
        print()

    if file_failures:
        print(f"ðŸ“ FILE ISSUE EXAMPLES:")
        for filename, reason in file_failures[:3]:
            print(f"  â€¢ {filename}: {reason}")
        if len(file_failures) > 3:
            print(f"  ... and {len(file_failures) - 3} more")
        print()

    if sheet_failures:
        print(f"ðŸ“‹ SHEET ISSUE EXAMPLES:")
        for filename, reason in sheet_failures[:3]:
            print(f"  â€¢ {filename}: {reason}")
        if len(sheet_failures) > 3:
            print(f"  ... and {len(sheet_failures) - 3} more")
        print()

    if data_failures:
        print(f"ðŸ“Š DATA ISSUE EXAMPLES:")
        for filename, reason in data_failures[:3]:
            print(f"  â€¢ {filename}: {reason}")
        if len(data_failures) > 3:
            print(f"  ... and {len(data_failures) - 3} more")
        print()

    print(f"="*80)
    print(f"")

    # â”€â”€â”€ PHASE 4: MATCH REPORTS TO TEMPLATES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    progress(65, "Matching reports to templates...")
    results = {}

    for report_data in report_data_list:
        print(f"ðŸ” Matching report '{report_data['filename']}'")
        print(f"    report_facility       = {report_data['report_facility']!r}")
        print(f"    matched_template_name = {report_data['matched_template_name']!r}")
        
        # Only show entries for that facility
        print("    RELEVANT template_entries:")
        relevant_entries = [e for e in template_entries if e["facility"] == report_data["matched_template_name"]]
        for e in relevant_entries:
            print(f"      â€¢ facility={e['facility']!r}, date={e['date']!r}")
        
        # Check available dates
        dates = [e["date"] for e in template_entries if e["facility"] == report_data["matched_template_name"]]
        print(f"    template dates for '{report_data['matched_template_name']}': {dates}")
        print(f"    report_date needed: {report_data['report_date']}")

        candidates = [
            e for e in template_entries
            if e["facility"] == report_data["matched_template_name"]
            and e["date"] == report_data["report_date"]
        ]
        
        if not candidates:
            skipped_reports.append((report_data["filename"], f"No matched date {report_data['report_date']}"))
            print("    âŒ No candidates, skipping\n")
            continue

        # Build results
        t = candidates[0]
        comparison_debug_log[t["facility"]]["Compared"] = True
        key = (t["facility"], report_data["report_date"])
        
        # Calculate actual HPPD values
        actual_hppd = report_data["actual_hours"] / t["census"] if t["census"] > 0 else 0
        actual_cna_hppd = report_data["actual_cna_hours"] / t["census"] if t["census"] > 0 else 0
        actual_rn_lpn_hppd = report_data["actual_rn_lpn_hours"] / t["census"] if t["census"] > 0 else 0

        results[key] = [
            {
                "Facility": t["facility"],
                "Type": "Projected",
                "Total HPPD": round(t["proj_total"], 2),
                "CNA HPPD": round(t["proj_cna"], 2),
                "RN+LPN HPPD": round(t["proj_nurse"], 2),
                "CNA Agency %": round(t["proj_agency_cna"], 2),
                "RN+LPN Agency %": round(t["proj_agency_nurse"], 2),
                "Total Agency %": round(t["proj_agency_total"], 2),
                "Notes": t.get("note", ""),
                "Date": report_data["report_date"]
            },
            {
                "Facility": t["facility"],
                "Type": "Actual",
                "Total HPPD": round(actual_hppd, 2),
                "CNA HPPD": round(actual_cna_hppd, 2),
                "RN+LPN HPPD": round(actual_rn_lpn_hppd, 2),
                "CNA Agency %": report_data["actual_agency_cna_pct"],
                "RN+LPN Agency %": report_data["actual_agency_nurse_pct"],
                "Total Agency %": report_data["actual_agency_total_pct"],
                "Notes": t.get("note", ""),
                "Date": report_data["report_date"]
            }
        ]
        print("    âœ… Matched and will be included\n")

    print(f"Generated results for {len(results)} facilities")

    # â”€â”€â”€ PHASE 5: EXCEL GENERATION â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    progress(80, "Generating Excel output...")
    
    # Pre-calculate all difference rows and column widths
    all_difference_rows = {}
    column_headers = [
        "Facility", "Type", "Total HPPD", "CNA HPPD", "RN+LPN HPPD",
        "CNA Agency %", "RN+LPN Agency %", "Total Agency %",
        "Notes", "Date"
    ]
    
    # Initialize column widths with header lengths
    column_widths = {header: len(header) for header in column_headers}
    
    for key in results.keys():
        projected_row = results[key][0]
        actual_row = results[key][1]
        
        difference_row = {"Type": "Difference", "Facility": "", "Date": projected_row["Date"]}
        for col_name in column_headers:
            if col_name in ("Facility", "Type", "Date"):
                continue
            proj_val = projected_row.get(col_name)
            act_val = actual_row.get(col_name)
            if isinstance(proj_val, (int, float)) and isinstance(act_val, (int, float)):
                difference_row[col_name] = round(proj_val - act_val, 2)
            else:
                difference_row[col_name] = None
        
        all_difference_rows[key] = difference_row
        
        # Calculate column widths
        for row_data in [projected_row, actual_row, difference_row]:
            for header in column_headers:
                if header == "Facility" and row_data["Type"] in ["Actual", "Difference"]:
                    content = ""
                else:
                    content = str(row_data.get(header, ""))
                
                content_width = len(content)
                column_widths[header] = max(column_widths[header], content_width)
    
    # Create output Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "HPPD Comparison"
    current_row = 1
    header_written = False

    def write_section(title, keys):
        nonlocal current_row, header_written
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=20)
        current_row += 1

        if not keys:
            ws.cell(row=current_row, column=1, value="No data available for this category.")
            current_row += 2
            return

        for col_idx, col_name in enumerate(column_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True, size=16)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if not header_written:
            ws.freeze_panes = ws.cell(row=current_row + 1, column=1)
            header_written = True

        current_row += 1

        for key in keys:
            projected_row = results[key][0]
            actual_row = results[key][1]
            difference_row = all_difference_rows[key]

            for row_data in [projected_row, actual_row, difference_row]:
                for col_idx, col_name in enumerate(column_headers, 1):
                    if col_name == "Facility" and row_data["Type"] in ["Actual", "Difference"]:
                        val = ""
                    else:
                        val = row_data.get(col_name, "")
                    
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = Font(size=14, italic=(row_data["Type"] == "Difference"))
                    
                    # Color coding for rows
                    if row_data["Type"] == "Projected":
                        cell.fill = PatternFill("solid", fgColor="D1CFCF")
                    elif row_data["Type"] == "Actual":
                        cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    elif row_data["Type"] == "Difference":
                        red_green_cols = (
                            "Total HPPD", "CNA HPPD", "RN+LPN HPPD",
                            "CNA Agency %", "RN+LPN Agency %", "Total Agency %"
                        )
                        if col_name in red_green_cols:
                            diff_val = difference_row.get(col_name)
                            if isinstance(diff_val, (int, float)):
                                if diff_val < 0:
                                    cell.fill = PatternFill("solid", fgColor="C8E6C9")
                                else:
                                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
                            else:
                                cell.fill = PatternFill("solid", fgColor="FFFACD")
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    
                    if col_name == "Date":
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        
                current_row += 1

        current_row += 2

    # Categorize results
    group1, group2, group3 = [], [], []
    for key, rows in results.items():
        actual = rows[1]
        hppd = actual["Total HPPD"]
        cna = actual["CNA HPPD"]
        rn = actual["RN+LPN HPPD"]
        if 3.0 <= hppd <= 3.3 and 2.00 <= cna <= 2.06 and rn <= 1.2:
            group1.append(key)
        elif 3.0 <= hppd <= 3.3 and (cna < 2.0 or rn > 1.2):
            group2.append(key)
        elif (hppd < 3.0 or hppd > 3.3) and (cna < 2.0 or rn > 1.2):
            group3.append(key)

    write_section("Good HPPD & Good Split (3.0<HPPD<3.3, 2.00<CNA<2.06, RN+LPN<=1.20)", group1)
    write_section("Good HPPD & Bad Split (3.0<HPPD<3.3, CNA<2.00, RN+LPN>1.20)", group2)
    write_section("Bad HPPD & Bad Split (HPPD>3.3 | HPPD<3.0, CNA<2.00, RN+LPN>1.20)", group3)

    # Set column widths
    for col_idx, header in enumerate(column_headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths[header] + 4

    # Add skipped templates sheet
    ws_skipped = wb.create_sheet(title="Skipped Templates")
    ws_skipped.append(["File Name", "Reason", "Category"])
    for filename, reason in skipped_templates:
        category = "Mac OS Hidden File" if "Mac OS hidden" in reason else "Invalid Data" if "Invalid" in reason else "File Error"
        ws_skipped.append([filename, reason, category])
    if not skipped_templates:
        ws_skipped.append(["âœ… No skipped templates", "", ""])
    ws_skipped.column_dimensions["A"].width = 40
    ws_skipped.column_dimensions["B"].width = 50
    ws_skipped.column_dimensions["C"].width = 20

    # Add skipped reports sheet
    ws_skipped_reports = wb.create_sheet(title="Skipped Reports")
    ws_skipped_reports.append(["File Name", "Reason", "Category"])
    for filename, reason in skipped_reports:
        category = "Mac OS Hidden File" if "Mac OS hidden" in reason else "Name Matching Issue" if "No matched facility" in reason else "File Error"
        ws_skipped_reports.append([filename, reason, category])
    if not skipped_reports:
        ws_skipped_reports.append(["âœ… No skipped reports", "", ""])
    ws_skipped_reports.column_dimensions["A"].width = 40
    ws_skipped_reports.column_dimensions["B"].width = 50
    ws_skipped_reports.column_dimensions["C"].width = 20

    # âœ… STEP 5: Write comparison debug log
    debug_df = pd.DataFrame.from_dict(comparison_debug_log, orient='index')
    debug_df.index.name = "Facility"
    debug_df.reset_index(inplace=True)

    ws_debug = wb.create_sheet(title="Comparison Debug Log")
    ws_debug.append(debug_df.columns.tolist())
    for row in debug_df.itertuples(index=False):
        ws_debug.append(list(row))

    # Optional: widen columns for clarity
    for col_idx, col_name in enumerate(debug_df.columns, 1):
        col_letter = get_column_letter(col_idx)
        ws_debug.column_dimensions[col_letter].width = max(15, len(col_name) + 4)

    # Save the file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_output_path = os.path.join(output_path, f"HPPD_Comparison_{timestamp}.xlsx")
    wb.save(final_output_path)
    
    progress(100, "âœ… Analysis complete!")
    print("Excel file created successfully!")
    return final_output_path